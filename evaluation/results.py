from evaluation.subject_classifier_metrics import precision, recall, f1
from tools.compute_subjects_list import compute_subjects_list
import pathlib
import os
from evaluation.rmse import RMSE
from evaluation.drmse import dRMSE
from evaluation.cg_ega.cg_ega import CG_EGA
import numpy as np
import openpyxl
from evaluation.r import r
from evaluation.fit import fit
from evaluation.time_lag import time_lag
from evaluation.mape import MAPE
import pickle
import misc
import pandas as pd


class ResultsTargetAnalyzer():
    """
        ResultsAnalyzer object that instiate results objects for a given population and compute the overall
        results.
    """

    def __init__(self, source_dataset, target_dataset, model_name, exp_name, freq=misc.freq):
        self.model_name = model_name
        self.freq = freq
        self.source_dataset = source_dataset
        self.target_dataset = target_dataset
        self.exp_name = exp_name

    def analyze(self, subject="all", details=False):
        """
            Compute the overall results of the given population
            :param subject: name of the subject, supports "all"
            :return: dict of mean results accros population, dict of std results accros population
        """
        datasets_subjects = compute_subjects_list(self.target_dataset, subject)

        res_list = []
        for dataset, subject in datasets_subjects:
            # load results
            res = ResultsTarget(self.exp_name, self.source_dataset, self.target_dataset, subject, self.model_name)

            # get and store results
            res_list.append(res.get_results())

        df = pd.DataFrame(res_list, columns=list(res_list[0].keys()))

        if details:
            return dict(df.mean()), dict(df.std()), df
        else:
            return dict(df.mean()), dict(df.std())


class ResultsTarget():
    """
        Results object that contains all the different metrics for on subject.
        Parameters:
            - model_name: name of the model (e.g., "ELM")
            - ph: prediction horizon in minutes (e.g., 30)
            - dataset: name of the dataset (e.g, "Ohio")
            - subject: name of the subject from the dataset (e.g., "591")
            - freq: sampling frequency in minutes (e.g., 5)
            - results: if not None, results are computed from the given array
            - file: if not None, results are computed from the given file, handles "auto"
    """

    def __init__(self, exp_name, source_dataset, target_dataset, target_subject, model_name, freq=misc.freq,
                 results=None, file="auto"):

        if results is not None:
            self.results = results
        else:
            if file == "auto":
                file = os.path.join(misc.path, "results", source_dataset + "_2_" + target_dataset,
                                    model_name + "_" + exp_name, target_dataset + target_subject + ".res")
            self.results = self.load(file)

        self.results_glucose = [split.transpose(2, 0, 1) for split in self.results]

        self.exp_name = exp_name
        self.freq = freq
        self.source_dataset = source_dataset
        self.target_dataset = target_dataset
        self.target_subject = target_subject
        self.model_name = model_name

        # if file is not None:
        #     if file == "auto":
        #         file_name = "ph" + str(ph) + "_" + dataset + subject + "_" + model_name + "_.res"
        #         self.load(os.path.join("results", "ph" + str(ph), model_name, file_name))
        #     else:
        #         self.load(file)
        #
        # elif results is not None:
        #     # reshape the results so that day have the shape (split, 2, day, val)
        #     self.results = [split.transpose(2, 0, 1) for split in results]
        #
        # self.model = model_name
        # self.freq = freq
        # self.ph = ph
        # self.dataset = dataset
        # self.subject = subject

    def get_results(self):
        """
            Compute the overall results, averaged for every split
            :return: dict with the following keys: RMSE, dRMSE, r, fit, TL, AP, BE, EP
        """
        # RMSE
        rmse = np.mean([RMSE(*res) for res in self.results_glucose])

        # RMSE
        mape = np.mean([MAPE(*res) for res in self.results_glucose])

        # dRMSE
        drmse = np.mean([dRMSE(*res, self.freq) for res in self.results_glucose])

        # CG-EGA
        cg_ega = np.mean([CG_EGA(*res, self.freq).reduced() for res in self.results_glucose], axis=0)

        r_metric = np.mean([r(*res) for res in self.results_glucose])
        fit_metric = np.mean([fit(*res) for res in self.results_glucose])
        time_lag_metric = np.mean([time_lag(*res, freq=self.freq) for res in self.results_glucose])

        return {
            "RMSE": rmse,
            "dRMSE": drmse,
            "MAPE": mape,
            "r": r_metric,
            "fit": fit_metric,
            "TL": time_lag_metric,
            "AP": cg_ega[0],
            "BE": cg_ega[1],
            "EP": cg_ega[2],
        }

    def plot(self, split=0, day=0):
        """
            Plot the results of a given day from a given split
            :param split: split number
            :param day: day number
            :return: /
        """
        y_true = self.results_glucose[split][0]
        y_pred = self.results_glucose[split][1]

        CG_EGA(y_true, y_pred, self.freq).plot(day=day)

    def to_excel(self, params, file_name="results.xlsx"):
        """
            Save the results into an excel file
            :param params:
            :param n_splits:
            :param file_name:
            :return:
        """
        file = os.path.join(misc.path, "results", file_name)

        results = self.get_results()
        misc_params = {"source_dataset": self.source_dataset, "target_dataset": self.target_dataset,
                       "target_subject": self.target_subject}

        data = {**misc_params, **params, **results}

        # if file not exist, create it with appropriate header
        if not pathlib.Path(file).is_file():
            wb = openpyxl.Workbook()
            wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
        else:
            wb = openpyxl.load_workbook(file)

        if not self.model_name in wb.sheetnames:
            wb.create_sheet(self.model_name)
            ws = wb[self.model_name]
            ws.append(list(data.keys()))
        else:
            ws = wb[self.model_name]

        ws.append(list(data.values()))
        wb.save(file)

    def save(self):
        """
            Save results object to file
            :param suffix: suffix to append to the file name
            :return: /
        """
        dir = os.path.join(misc.path, "results", self.source_dataset + "_2_" + self.target_dataset,
                           self.model_name + "_" + self.exp_name)
        file_name = self.target_dataset + self.target_subject + ".res"

        if not os.path.exists(dir): os.makedirs(dir)
        file = os.path.join(dir, file_name)

        with open(file, "wb") as the_file:
            pickle.dump(self.results, the_file)

    def load(self, file):
        """
            load Results object from file
            :param file: file
            :return: /
        """
        with open(file, "rb") as the_file:
            return pickle.load(the_file)

    def save_day(self, split, day):
        """
            Save on particular day (ground_truth, predictions, AP, BE, EP)
            :param split: split number
            :param day: day number
            :return: /
        """
        y_true = self.results_glucose[split][0, day, :].reshape(1, -1)
        y_pred = self.results_glucose[split][1, day, :].reshape(1, -1)

        df = CG_EGA(y_true, y_pred, self.freq).per_sample()
        ap = df[df["CG_EGA"] == "AP"]
        be = df[df["CG_EGA"] == "BE"]
        ep = df[df["CG_EGA"] == "EP"]

        dir = "s" + str(split) + "_d" + str(day)
        path = os.path.join(os.path.dirname(__file__), "..", "results", self.model_name, dir)
        if not os.path.exists(path): os.makedirs(path)

        df.loc[:, ["time", "y_true"]].to_csv(os.path.join(path, "y_true"), sep=" ", index=False)
        df.loc[:, ["time", "y_pred"]].to_csv(os.path.join(path, "y_pred"), sep=" ", index=False)
        ap.loc[:, ["time", "y_pred"]].to_csv(os.path.join(path, "time_ap"), sep=" ", index=False)
        be.loc[:, ["time", "y_pred"]].to_csv(os.path.join(path, "time_be"), sep=" ", index=False)
        ep.loc[:, ["time", "y_pred"]].to_csv(os.path.join(path, "time_ep"), sep=" ", index=False)
        ap.loc[:, ["y_true", "y_pred"]].to_csv(os.path.join(path, "p_ega_ap"), sep=" ", index=False)
        be.loc[:, ["y_true", "y_pred"]].to_csv(os.path.join(path, "p_ega_be"), sep=" ", index=False)
        ep.loc[:, ["y_true", "y_pred"]].to_csv(os.path.join(path, "p_ega_ep"), sep=" ", index=False)
        ap.loc[:, ["dy_true", "dy_pred"]].to_csv(os.path.join(path, "r_ega_ap"), sep=" ", index=False)
        be.loc[:, ["dy_true", "dy_pred"]].to_csv(os.path.join(path, "r_ega_be"), sep=" ", index=False)
        ep.loc[:, ["dy_true", "dy_pred"]].to_csv(os.path.join(path, "r_ega_ep"), sep=" ", index=False)


class ResultsSource():
    def __init__(self, exp_name, source_dataset, target_dataset, target_subject, model_name, freq=misc.freq,
                 results=None,
                 file="auto"):
        if results is not None:
            self.results = results
        else:
            if file == "auto":
                file = os.path.join(misc.path, "results", source_dataset + "_2_" + target_dataset,
                                    model_name + "_" + exp_name, target_dataset + target_subject + ".res")
            self.results = self.load(file)

        # (split, glucose+subject, day, val, y_true+y_pred) =>
        # glucose: (split, y_true+y_pred, day, val) + subject: (split, y_true+y_pred, day, val)
        self.results_glucose = [np.rollaxis(split[0], 2, 0) for split in self.results]
        self.results_subjects = [np.rollaxis(split[1], 2, 0) for split in self.results]

        self.exp_name = exp_name
        self.freq = freq
        self.source_dataset = source_dataset
        self.target_dataset = target_dataset
        self.target_subject = target_subject
        self.model_name = model_name

    def get_results(self):
        """
            Compute the overall results, averaged for every split
            :return: dict with the following keys: RMSE, dRMSE, r, fit, TL, AP, BE, EP
        """
        # RMSE
        rmse = np.mean([RMSE(*res) for res in self.results_glucose])

        # RMSE
        mape = np.mean([MAPE(*res) for res in self.results_glucose])

        # dRMSE
        drmse = np.mean([dRMSE(*res, self.freq) for res in self.results_glucose])

        # CG-EGA
        cg_ega = np.mean([CG_EGA(*res, self.freq).reduced() for res in self.results_glucose], axis=0)

        r_metric = np.mean([r(*res) for res in self.results_glucose])
        fit_metric = np.mean([fit(*res) for res in self.results_glucose])
        time_lag_metric = np.mean([time_lag(*res, freq=self.freq) for res in self.results_glucose])

        precision_metric = np.mean([precision(*res) for res in self.results_subjects])
        recall_metric = np.mean([recall(*res) for res in self.results_subjects])
        f1_metric = np.mean([f1(*res) for res in self.results_subjects])

        return {
            "RMSE": rmse,
            "dRMSE": drmse,
            "MAPE": mape,
            "r": r_metric,
            "fit": fit_metric,
            "TL": time_lag_metric,
            "AP": cg_ega[0],
            "BE": cg_ega[1],
            "EP": cg_ega[2],
            "precision": precision_metric,
            "recall": recall_metric,
            "f1": f1_metric,
        }

    def save(self):
        """
            Save results object to file
            :param suffix: suffix to append to the file name
            :return: /
        """
        dir = os.path.join(misc.path, "results", self.source_dataset + "_2_" + self.target_dataset,
                           self.model_name + "_" + self.exp_name)
        file_name = self.target_dataset + self.target_subject + ".res"

        if not os.path.exists(dir): os.makedirs(dir)
        file = os.path.join(dir, file_name)

        with open(file, "wb") as the_file:
            pickle.dump(self.results, the_file)

    def load(self, file):
        """
            load Results object from file
            :param file: file
            :return: /
        """
        with open(file, "rb") as the_file:
            return pickle.load(the_file)
